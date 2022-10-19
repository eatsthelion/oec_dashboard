import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import string 
from datetime import datetime
from Backend.database import DOCCATALOG, PROJECTDB, DB_connect, DATABASE, ITEMCATALOG, DB_database_find_replace

EMPTYLIST = ["None", None, 'nan', 'NaN','']
ENTRYLIMIT = 10000

#file = r'D:\My Drive\MASTER OEC Project Database 081622.xlsx'
#
#csvfile = pd.read_excel(file)
#
insertsql = """INSERT INTO project_info VALUES 
    ('{}','{}','{}','{}','{}',
    '{}','{}','{}','{}','{}',
    '{}','{}','{}','{}','{}',
    '{}','{}')"""
#
project_dict = {}
projects = DB_connect("SELECT rowid, oec_job, client_job FROM project_info", database=PROJECTDB)
for project in projects: project_dict[project[1]+" "+str(project[2])]=project[0]

def excel_input(csvfile, entry_function):
    entry_array = []
    for index in range(2,len(csvfile)):
        entry = entry_function(index, csvfile)
        if not entry: continue
        entry_array.append(entry)
        
        if len(entry_array)<ENTRYLIMIT: continue
        DB_connect(entry_array,database=PROJECTDB)
        for en in entry_array: print(en)
        entry_array=[]
    DB_connect(entry_array,database=PROJECTDB)
    for en in entry_array: print(en)

def project_insert(index,csvfile):
    active_status = csvfile.iloc[index,0]
    if active_status == "Y": active_status == 'ACTIVE'
    elif active_status == "N": active_status == 'COMPLETED'
    comments = str(csvfile.iloc[index,7])
    if "CANCEL" in comments: active_status='CANCELLED'
    cwa_type = csvfile.iloc[index,3]
    if cwa_type not in ['LS','TM']: cwa_type = ''
    client = str(csvfile.iloc[index,1]).upper()
    newentry = insertsql.format(
        str(csvfile.iloc[index,2]).replace("'", "''").strip(),                              #oec_job
        str(csvfile.iloc[index,4]).replace("'", "''").strip(),                              #client_job
        str(client).replace("'", "''").strip(),                                             #client
        active_status,                                                                      #active_status
        str(csvfile.iloc[index,6]).replace("'", "''").strip(),                              #project_name
        str(csvfile.iloc[index,5]).replace("'", "''").strip(),                              #location
        str(csvfile.iloc[index,8]).replace("'", "''").strip(),                              #project_engineer
        str(csvfile.iloc[index,9]).replace("'", "''").strip(),                              #outdoor_designers
        str(csvfile.iloc[index,10]).replace("'", "''").strip(),                             #indoor_designers
        '',                                                                                 #project_type
        cwa_type,                                                                           #cwa_type
        str(csvfile.iloc[index,16]).replace("'", "''").strip(),                             #current_stage
        str(csvfile.iloc[index,15]).replace("'", "''").strip(),                                                                                 #current_phase
        csvfile.iloc[index,17],                                                             #current_percent_complete
        '',                                                                                 #creation_date
        '',                                                                                 #modify_date
        ''                                                                                  #complete_date
        )
    return newentry

def change_log_insert(index):
    pass

def update_create_dates(index,csvfile):
    oec_job = str(csvfile.iloc[index,2])

    created_date = datetime.today()
    modify_date = datetime(year=1986,month=8,day=26)
    for column in range(len(csvfile.columns)):
        try:created_date = min(csvfile.iloc[index,column],created_date)
        except: pass
        try:modify_date =  max(csvfile.iloc[index,column],modify_date)
        except: pass
    #print(oec_job, rowid, 'Create Date:', created_date, 'Modify Date:', modify_date)
    new_created_date = created_date.strftime('%Y-%m-%d %H:%M:%S')
    new_modify_date = modify_date.strftime('%Y-%m-%d %H:%M:%S')
    if '2022-09-13' in new_created_date : 
        try: 
            new_created_date = datetime(year=int('20'+oec_job[:2]),month=1,day=1)
            new_created_date = new_created_date.strftime('%Y-%m-%d %H:%M:%S')
        except: pass
    if '1986-08-26' in new_modify_date: new_modify_date = new_created_date
    return "UPDATE project_info SET creation_date = '{}', modify_date = '{}' WHERE oec_job = '{}'".format(new_created_date, new_modify_date, oec_job,)

def insert_dates(index,csvfile):
    oec_job = csvfile.iloc[index,2]
    client_job = csvfile.iloc[index,4]
    try: rowid = project_dict[oec_job+' '+str(client_job)]
    except: 
        print('No Project Match')
        return False
    insert_str = 'INSERT INTO project_dates VALUES '
    for column in range(32, len(csvfile.columns)):
        cell = csvfile.iloc[index,column]
        milestone = csvfile.iloc[1,column]
        if milestone in EMPTYLIST: continue
        forecast_date = ''
        actual_date = ''

        try: 
            entry_date = cell.strftime('%Y-%m-%d %H:%M:%S')
            if str(csvfile.iloc[index,column+1]).upper() == 'F': forecast_date = entry_date
            else: actual_date = entry_date
        except: continue

        insert_str += "('{}','{}','{}','{}','',''), ".format(rowid, milestone, forecast_date,actual_date)
    if insert_str == 'INSERT INTO project_dates VALUES ': 
        print("{} has no dates.".format(oec_job))
        return False
    
    print('Dates for {} have been found!'.format(oec_job))
    return insert_str.strip(', ')

def insert_budget_info(index, csvfile):
    oec_job = csvfile.iloc[index,2]
    client_job = csvfile.iloc[index,4]
    try: rowid = project_dict[oec_job+' '+str(client_job)]
    except: return False

def document_projects():
    projects = DB_connect("SELECT rowid, client_job FROM project_info WHERE NOT (client_job = '' or client_job = 'None' or client_job = 'nan')", database=PROJECTDB)
    project_rowid_dict = {}
    print(projects)
    for project in projects: 
        project_rowid_dict[project[1]] = project[0]
    documents = DB_connect("SELECT rowid, client_job FROM documents WHERE NOT (client_job = '' or client_job = 'None' or client_job = 'nan')", database=DOCCATALOG)

    entry_array = []
    for document in documents:
        try: 
            project_id = project_rowid_dict[document[1]]
        except: continue
        entry_array.append("UPDATE documents SET project_id = '{}' WHERE rowid = '{}'".format(project_id,document[0]))
        
        if len(entry_array)<ENTRYLIMIT: continue
        DB_connect(entry_array,database=DOCCATALOG)
        for en in entry_array: print(en)
        entry_array=[]
    DB_connect(entry_array,database=DOCCATALOG)
    for en in entry_array: print(en)

def apply_project_id_to_budget():
    update_list = []
    purchase_orders = DB_connect("SELECT rowid, oec_job, client_job, location FROM project_budget", database=PROJECTDB)
    for po in purchase_orders:
        projects = DB_connect("""
            SELECT rowid 
            FROM project_info 
            WHERE oec_job LIKE '%{}%' 
            AND client_job='{}' 
            AND location='{}'
            """.format(
            po[1],po[2],po[3]), database=PROJECTDB)

        if len(projects) == 0:
            projects = DB_connect("""
                SELECT rowid 
                FROM project_info 
                WHERE oec_job LIKE '%{}%'
                AND client_job='{}'
                """.format(
                po[1],po[2]), database=PROJECTDB)

        if len(projects) == 0:
            projects = DB_connect("""
                SELECT rowid 
                FROM project_info 
                WHERE oec_job LIKE '%{}%'
                """.format(
                po[1]), database=PROJECTDB)
        if len(projects) == 0: continue
        print(po[1], projects)
        update_list.append("""UPDATE project_budget SET project_id ='{}' WHERE rowid={}""".format(projects[0][0],po[0]))

    DB_connect(update_list, database=PROJECTDB)

def recode_oec_projects(excel_file):
    dataframe = pd.read_excel(excel_file, dtype=str)
    project_dict = {}
    for row in range(1,len(dataframe.index)):
        oec_num = str(dataframe.iloc[row,2]).strip()
        if oec_num in EMPTYLIST: continue
        if oec_num not in project_dict:
            project_dict[oec_num] = 1
        else:
            project_dict[oec_num] += 1

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    for row in range(3,sheet.max_row+1):
        cur_cell = sheet.cell(row,3)
        oec_num = str(cur_cell.value).strip()
        if oec_num in EMPTYLIST: continue
        if project_dict[oec_num]==1: continue
        
        for loop_count in range(1,project_dict[oec_num]+1):
            new_oec_num = oec_num+get_column_letter(loop_count)
            if new_oec_num in project_dict: continue
            cur_cell.value = new_oec_num
            project_dict[new_oec_num] = 1
            break
                
    filenames = os.path.splitext(excel_file)
    outputfile = filenames[0]+' Copy'+filenames[1]
    wb.save(outputfile)
    wb.close()
    os.startfile(outputfile)

    return outputfile

def project_database_initialize():
    print('formatting excel files...')
    outputfile = r'G:\My Drive\MASTER OEC Project Database 090122 Copy.xlsx'
    
    csvfile = pd.read_excel(outputfile)
    print("inserting projects into database...")
    excel_input(csvfile, project_insert)
    project_dict = {}
    projects = DB_connect("SELECT rowid, oec_job FROM project_info", database=PROJECTDB)
    for project in projects: project_dict[project[1]]=project[0]
    print("updating create dates...")
    excel_input(csvfile, update_create_dates)

def project_comments_initialize():
    pd_dataframe1 = pd.read_excel(r'D:\My Drive\MASTER OEC Project Database 090122 Copy.xlsx')
    project_comment_dict = {}

    for row in range(len(pd_dataframe1.index)):
        oec_num = pd_dataframe1.iloc[row, 2]
        comments = pd_dataframe1.iloc[row, 7]
        print(comments)
        project_comment_dict[oec_num] = comments

    workbook = openpyxl.load_workbook(r'D:\My Drive\OEC Project Catalog Old.xlsx')
    sheet = workbook.active

    for row in range(1,sheet.max_row+1):
        if sheet.cell(row, 1).value not in project_comment_dict: continue
        sheet.cell(row, 7).value = project_comment_dict[sheet.cell(row, 1).value]

    workbook.save(r'D:\My Drive\OEC Project Catalog New Copy.xlsx')
    workbook.close()

if __name__ == '__main__':
    recode_oec_projects(r'G:\PROJECT CONTROL\MASTER OEC Project Database 091522.xlsx')



#SELECT 
#p.oec_job, p.client_job, p.client, p.active_status, p.project_name, 
#p.location, p.project_engineer, p.current_stage, p.current_phase, p.current_percent_complete, 
#pb.purchase_order, pb.cwa_num, pb.cwa_type, pb.cwa_recieved_amount, pb.current_balance
#FROM project_info AS p
#JOIN (
#	SELECT MAX(rowid) as max_id, project_id
#	FROM project_budget 
#	GROUP BY project_id
#) AS pb_max ON (pb_max.project_id = p.rowid)
#JOIN project_budget AS pb ON pb.rowid = pb_max.max_id
#WHERE active_status = 'ACTIVE'
#ORDER BY p.creation_date DESC
    

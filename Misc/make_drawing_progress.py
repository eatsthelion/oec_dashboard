#############################################################################################
# make_drawing_progress.py
# 
# Created: 9/02/22
# Creator: Ethan de Leon
# Purposes:
#   - Create a progress drawing list with titles 
#   - Update an existing progress drawing list with the latest document addresses and 
#     revisions
# Required Installs: pandas
#############################################################################################

import os
import openpyxl
import pandas as pd
import tkinter as tk

from tkinter import *
from tkinter import filedialog, messagebox 
from datetime import datetime

from GUI.GUI_Mains import FONT, FONTBOLD, OECCOLOR
from GUI.window_main import ProgramWindow
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from GUI.widgets.terminal import Terminal
from Backend.exports import export_init
from Backend.database import DB_connect
from Backend.path_analyzer import PathAnalyzer
from Misc.yearly_archiver import AutoArchiver

PROGRAMTITLE = 'Drawing Progress Report'
WIDGETWIDTH = 60

PROGRAMFOLDER = export_init(PROGRAMTITLE)

class DWGProgressReportMkr(ProgramWindow):
    def __init__(self) -> None:
        super().__init__(configure=self.configure)

    def configure(self):
        self.settings()
        self.widgets()
        self.root.update()
        self.root.minsize(self.root.winfo_width(),self.root.winfo_height())
        self.root.geometry('+{}+{}'.format(int((self.root.winfo_screenwidth()-self.root.winfo_width())/2),
            int((self.root.winfo_screenheight()-self.root.winfo_height())/2)))
        self.root.resizable(0,0)

    def settings(self):
        self.root.title(PROGRAMTITLE)

    def widgets(self):
        titleframe = tk.Frame(self.mainframe,bg='white')
        titlelabel = tk.Label(titleframe, font = ("montserrat extrabold",24,'bold'), 
            text="★ "+PROGRAMTITLE.upper()+" ★", bg=titleframe.cget('background'),
            height=2)
        
        titleframe.pack(fill='x')
        titlelabel.pack()

        self.frame = tk.LabelFrame(self.mainframe, bg=OECCOLOR, text='Create D.P.R From Files', font=FONTBOLD, fg='white')
        self.button_frame = tk.Frame(self.frame,bg=self.frame.cget('background'))
        self.substation_label   = tk.Label(self.frame, bg=self.frame.cget('background'),fg='white', text='Location*:', font=FONT)
        self.client_order_label = tk.Label(self.frame, bg=self.frame.cget('background'),fg='white', text='Client Order:', font=FONT)
        self.oec_number_label   = tk.Label(self.frame, bg=self.frame.cget('background'),fg='white', text='OEC Job #:', font=FONT)

        self.files_button       = tk.Button(self.button_frame, text='Select Files', font=FONT, command=self.get_files)
        self.clear_button       = tk.Button(self.button_frame, text='Clear Files', font=FONT, command=self.clear_files)

        self.substation_entry   = tk.Entry(self.frame, font=FONT, width=WIDGETWIDTH)
        self.client_order_entry = tk.Entry(self.frame, font=FONT, width=WIDGETWIDTH)
        self.oec_number_entry   = tk.Entry(self.frame, font=FONT, width=WIDGETWIDTH)
        self.files_entry        = tk.Text(self.frame, font=FONT, state=DISABLED, width=WIDGETWIDTH, height=10)
        self.start_button       = tk.Button(self.frame, text = '★ MAKE DWG PROGRESS REPORT ★', font = FONT, command=self.make_progress_report)

        self.update_frame = tk.LabelFrame(self.mainframe, bg=OECCOLOR, text='Update Existing D.P.R', font=FONTBOLD, fg='white')
        self.update_button      = tk.Button(self.update_frame,text = '★ UPDATE DWG PROGRESS REPORT ★', font = FONT)
        
        self.terminal = Terminal(self.mainframe)

        self.frame.pack(padx=10)
        self.update_frame.pack(padx=10, pady=(0,10), fill =BOTH)
        self.substation_label   .grid(row=0,column=0, padx=5, pady=(5,0),sticky=NW)
        self.client_order_label .grid(row=1,column=0, padx=5, pady=(5,0),sticky=NW)
        self.oec_number_label   .grid(row=2,column=0, padx=5, pady=(5,0),sticky=NW)
        self.button_frame       .grid(row=3,column=0, padx=5, pady=(5,0),sticky=NW)

        self.files_button       .grid(row=0,column=0, pady=(0,5),sticky=EW)
        self.clear_button       .grid(row=1,column=0, pady=(0,5),sticky=EW)

        self.substation_entry   .grid(row=0,column=1, padx=5, pady=(5,0),sticky=NW)
        self.client_order_entry .grid(row=1,column=1, padx=5, pady=(5,0),sticky=NW)
        self.oec_number_entry   .grid(row=2,column=1, padx=5, pady=(5,0),sticky=NW)
        self.files_entry        .grid(row=3,column=1, padx=5, pady=(5,0),sticky=NW)

        self.start_button       .grid(row=4, column=0, columnspan=2, sticky=NSEW,padx=5,pady=5)

        self.update_button      .pack(fill=BOTH, padx=5,pady=5)

    def get_files(self):
        self.files = filedialog.askopenfilenames()
        if not self.files: 
            return False
        self.files = sorted(self.files)
        
        insert_text = ''
        for file in self.files:insert_text+='{}. '.format(self.files.index(file)+1)+os.path.basename(file)+'\n'
        self.files_entry.configure(state=NORMAL)
        self.files_entry.delete('1.0',END)
        self.files_entry.insert('1.0',insert_text.strip('\n'))
        self.files_entry.configure(state=DISABLED)
            
    def make_progress_report(self):
        substation = self.substation_entry.get()
        if (not substation) or (not self.files): 
            messagebox.showerror('Missing Info', 'Please make sure that all information is provided.')
            return False
        today = datetime.today()
        outputfile = os.path.join(PROGRAMFOLDER, "{} Drawing Progress Report {}.xlsx".format(substation, today.strftime('%m-%d-%Y')))

        try: 
            if os.path.exists(outputfile): os.remove(outputfile)
        except:
            messagebox.showerror('File is open', '{} is open. Please close the file and try again.'.format(os.path.basename(outputfile)))
            return False

        client_order = self.client_order_entry.get()
        if not client_order: client_order = "_______"

        oec_number = self.oec_number_entry.get()
        if not oec_number: oec_number = "_______"

        self.terminal.show_terminal()
        self.terminal.print_terminal("Connecting to Document Archive...")

        # initializes the connection to the document archive
        self.archive = AutoArchiver()
        self.archive.yearly_database_initialize()

        self.terminal.print_terminal("Reading Header Columns...")

        pandas_dict = {'DWG #':[], 'REV':[], 'SH #':[], 'TITLE':[], 'PROGRESS':[], 'COMPLETE DATE':[], 'REMARKS':[]}
        location_dict = {"DWG #":[], 'REV':[],'LATEST ARCHIVED DRAWING':[]}
        for file in self.files:
            filename = os.path.basename(file)
            dwg_num = PathAnalyzer.findDwgNumPath(filename)
            self.terminal.print_terminal("Finding Title for Drawing {}...".format(dwg_num))
            rev_from_file = PathAnalyzer.findRevPath(filename)
            title, rev_from_archive, archive_location = self.findTitle(dwg_num)
            if rev_from_archive>rev_from_file:
                rev = rev_from_archive
            else: 
                rev = rev_from_file

            location_dict['DWG #'].append(dwg_num)
            location_dict['REV'].append(rev_from_archive)
            location_dict['LATEST ARCHIVED DRAWING'].append(archive_location)
            
            pandas_dict['DWG #'].append(dwg_num)
            pandas_dict['REV'].append(rev)
            pandas_dict['SH #'].append(PathAnalyzer.findSheetPath(filename))
            pandas_dict['TITLE'].append(title)
            pandas_dict['PROGRESS'].append(0)
            pandas_dict['COMPLETE DATE'].append('')
            pandas_dict['REMARKS'].append('')

        self.terminal.print_terminal("Exporting Data to Drawing List...")
        pd_dataframe = pd.DataFrame(pandas_dict)
        location_dataframe = pd.DataFrame(location_dict)
        writer = pd.ExcelWriter(outputfile, engine = 'xlsxwriter')
        pd_dataframe.to_excel(writer, sheet_name = 'Progress')
        location_dataframe.to_excel(writer, sheet_name = 'Previous Rev Files')
        writer.save()
        
        # Formating Workbook
        self.terminal.print_terminal("Formatting Drawing List...")
        workbook= openpyxl.load_workbook(outputfile)
        sheet = workbook['Progress'] 

        # sets Header
        sheet.oddHeader.center.text = '{} SUBSTATION\nDRAWING PROGRESS REPORT'.format(substation.upper())
        sheet.oddHeader.center.size = 14
        sheet.oddHeader.center.font = "Arial,Bold"
        sheet.oddHeader.right.text = 'ORDER #: {}\nOEC JOB #: {}'.format(client_order, oec_number)
        sheet.oddHeader.right.size = 14
        sheet.oddHeader.right.font = "Arial,Bold"
        sheet.oddHeader.left.text = 'AS OF {}'.format(today.strftime('%m/%d/%y'))
        sheet.oddHeader.left.size = 14
        sheet.oddHeader.left.font = "Arial,Bold"        

        sheet.delete_cols(1)
        #sheet.freeze_panes=sheet['A2']
        sheet.row_dimensions[1].height = 40
        # sets column widths
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 7
        sheet.column_dimensions['C'].width = 7
        sheet.column_dimensions['D'].width = 80
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions["F"].width = 22
        sheet.column_dimensions["G"].width = 30

        # printing formats
        openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(sheet, paper_size = 1, orientation='landscape')
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToHeight = False
        sheet.print_title_rows = '1:1' 

        # initializes cell format variables
        cell_font = Font(name='Arial', size='14')
        bold_font = Font(name='Arial', size='12')
        cell_boarder = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        greyFill = PatternFill(start_color='FFC1CDCD',
                   end_color='FFC1CDCD',
                   fill_type='solid')

        for col in range(1,sheet.max_column+1): 
            column = sheet.column_dimensions[get_column_letter(col)]
            column.border = cell_boarder
            column.font = cell_font

            sheet.cell(1,col).fill=greyFill
            sheet.cell(1,col).font=bold_font
            sheet.cell(1,col).alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for row in range(2,sheet.max_row+1):
            for col in range(1,sheet.max_column+1):
                sheet.cell(row,col).border = cell_boarder
                sheet.cell(row,col).font = cell_font

        #sheet.sheet_view.view = "pageLayout"

        workbook.save(outputfile)
        workbook.close()

        os.startfile(outputfile)
        self.clear_all()
        self.terminal.hide_terminal()

    def update_progress_report(self):

        pass

    def findTitle(self, dwg_num):
        select_query = "SELECT title, revision, file_address FROM documents WHERE dwg_num = '{}' ORDER BY revision DESC".format(dwg_num)

        for year in sorted(self.archive.yearly_database_dict):
            results = DB_connect(select_query, database=self.archive.yearly_database_dict[year]['database'])
            if len(results) == 0: continue
            return str(results[0][0]), str(results[0][1]), str(results[0][2])
        return '', '', ''

    def clear_files(self):
        self.files = None
        self.files_entry.configure(state=NORMAL)
        self.files_entry.delete('1.0',END)
        self.files_entry.configure(state=DISABLED)
        pass

    def clear_all(self):
        self.substation_entry.delete(0,END)
        self.client_order_entry.delete(0,END)
        self.oec_number_entry.delete(0,END)
        self.clear_files()

DWGProgressReportMkr()
# region Imports
import os
import openpyxl
import pandas as pd
import tkinter as tk

from copy import copy
from openpyxl.utils import get_column_letter
from tkinter import CENTER, DISABLED, END, EW, NSEW, W, messagebox, filedialog, NORMAL

from GUI.GUI_Mains import FONT
from GUI.widgets.terminal import Terminal
from GUI.window_main import PopupWindow

# endregion
# region MACROS

PROGRAMTITLE = 'Excel Table Comparison'
# endregion

class ExcelComparer(PopupWindow):
    def __init__(self, master, **kw) -> None:
        self.previous_file = None
        self.present_file = None
        super().__init__(master, width=1000, height=400, **kw)
        
    def configure(self):
        self.titlelabel.configure(text="★ "+PROGRAMTITLE.upper()+" ★")
        body_frame = tk.Frame(self.frame, bg=self.frame.cget('background'))
        self.previous_log_label = tk.Label(body_frame, font=FONT,fg='white', 
            bg=body_frame.cget('background'), text='Old File:')
        self.present_log_label  = tk.Label(body_frame, font=FONT,fg='white', 
            bg=body_frame.cget('background'), text='New File:')

        self.previous_log_entry = tk.Entry(body_frame, font=FONT, 
            state=DISABLED, width=50, disabledforeground='black', 
            disabledbackground='gray85')
        self.present_log_entry  = tk.Entry(body_frame, font=FONT, 
            state=DISABLED, width=50, disabledforeground='black', 
            disabledbackground='gray85')

        self.prev_button    = tk.Button(body_frame, font=FONT, text='SELECT', 
            command=lambda m='previous': self.file_getter(m))
        self.present_button = tk.Button(body_frame, font=FONT, text='SELECT', 
            command=lambda m='present': self.file_getter(m))
        
        starter_row_label = tk.Label(body_frame, font=FONT,fg='white', 
            bg=body_frame.cget('background'), text='Starting Row:')
        self.starter_row_entry = tk.Entry(body_frame, font=FONT, width=50)

        prime_col_label = tk.Label(body_frame, font=FONT,fg='white', 
            bg=body_frame.cget('background'), text='Primary Column:')
        self.prime_col_entry = tk.Entry(body_frame, font=FONT, width=50)
        
        self.execute_button = tk.Button(body_frame, font=FONT, 
            text='START COMPARISON', command=self.start_comparison)

        self.terminal = Terminal(self.frame, self)
        body_frame.pack(padx=10, pady=5)

        self.previous_log_label.grid(row=0 ,column=0 , padx=5, pady=5, sticky=W)
        self.present_log_label .grid(row=1 ,column=0 , padx=5, pady=5, sticky=W)

        self.prev_button   .grid(row=0, column = 2, padx=5, pady=5, sticky=EW)
        self.present_button.grid(row=1, column = 2, padx=5, pady=5, sticky=EW)

        self.previous_log_entry.grid(row=0 ,column=1 , padx=5, pady=5, sticky=NSEW)
        self.present_log_entry .grid(row=1 ,column=1 , padx=5, pady=5, sticky=NSEW)

        starter_row_label .grid(row=2, column=0, pady=5, padx=5, sticky=W)
        self.starter_row_entry .grid(row=2, column=1, pady=5, padx=5, sticky=EW)
        prime_col_label .grid(row=3, column=0, pady=5, padx=5, sticky=W)
        self.prime_col_entry .grid(row=3, column=1, pady=5, padx=5, sticky=EW)

        self.execute_button.grid(row=4, column=0, columnspan=3, padx=5, pady=5, sticky=EW)
        return super().configure()

    def file_getter(self, filetype='present'):
        file = filedialog.askopenfilename(title="Select a Spreadsheet", filetypes=[('Excel', '*.xls*')])
        if not file: 
            return False

        if filetype == 'present':
            self.present_file = file
        elif filetype == 'previous':
            self.previous_file = file

        filedict = {'present':self.present_log_entry, 'previous':self.previous_log_entry}
        filedict[filetype].configure(state=NORMAL)
        filedict[filetype].delete(0,END)
        filedict[filetype].insert(0,file)
        filedict[filetype].configure(state=DISABLED)

    def start_comparison(self):
        # Input Error Check and Dataframe initialization
        self.starting_row = self.starter_row_entry.get()
        self.terminal.show_terminal()
        self.terminal.print_terminal('Validating inputs...')
        if self.starting_row != '':
            try:
                self.starting_row = int(self.starting_row)
                if self.starting_row<=1: self.starting_row = ''
            except:
                messagebox.showerror('Starting Row Error', "The starting row input not valid.")
                self.terminal.hide_terminal()
                return False
        else:
            self.starting_row = 1
        try: 
            self.terminal.print_terminal('Reading Old File...')
            old_df = pd.read_excel(self.previous_file,dtype=str)
            if (self.starting_row != '')or(self.starting_row>=2):
                old_df.columns = old_df.iloc[self.starting_row-2]
                #old_df = old_df[self.starting_row+1:]
                old_df = old_df.drop(range(self.starting_row-1),axis=0)
                old_df.head()

            column_names = old_df.columns.values
            col_list = []
            for col in range(len(column_names)):
                if column_names[col] not in col_list:
                    col_list.append(column_names[col])
                    continue

                col_count = 1
                col_name = str(column_names[col]) + " ({})".format(col_count)
                while col_name in col_list:
                    col_count += 1
                    col_name = str(column_names[col]) + " ({})".format(col_count)
                col_list.append(col_name)
                column_names[col] = col_name

            old_df.columns = column_names

        except:
            messagebox.showerror('File Cannot Be Read', "The Previous Log File cannot be read.")
            self.terminal.hide_terminal()
            return False
        try: 
            self.terminal.print_terminal('Reading New File...')
            new_df = pd.read_excel(self.present_file,dtype=str)
            if (self.starting_row != '')or(self.starting_row>=2):
                new_df.columns = new_df.iloc[self.starting_row-2]
                #new_df = new_df[self.starting_row+1:]
                new_df = new_df.drop(range(self.starting_row-1),axis=0)
                new_df.head()
                
            column_names = new_df.columns.values
            col_list = []
            for col in range(len(column_names)):
                if column_names[col] not in col_list:
                    col_list.append(column_names[col])
                    continue

                col_count = 1
                col_name = str(column_names[col]) + " ({})".format(col_count)
                while col_name in col_list:
                    col_count += 1
                    col_name = str(column_names[col]) + " ({})".format(col_count)
                col_list.append(col_name)
                column_names[col] = col_name

            new_df.columns = column_names

        except:
            messagebox.showerror('File Cannot Be Read', "The Present Log File cannot be read.")
            self.terminal.hide_terminal()
            return False

        self.pivot_column = self.prime_col_entry.get().strip()
        if self.pivot_column not in column_names:
            messagebox.showerror('Pivot Column Error', "The Column \"{}\" does not exist in the selected files.".format(self.pivot_column))
            self.terminal.hide_terminal()
            return False

        
        
        self.compare_excel_changes(old_df, new_df)    
        return True
        #except:
        #    messagebox.showerror('ERROR', 'An error occured.')
        #    self.terminal.hide_terminal()
        #    return False

    def compare_excel_changes(self, old_df, new_df):
        # Finds all the format setting in each column of the old file
        self.terminal.print_terminal('Finding columns...')
        old_file = openpyxl.load_workbook(self.previous_file)
        sheet = old_file.active
        columns = []
        attribute_dict = {}

        for col in range(1, sheet.max_column+1):
            cur_cell = sheet.cell(self.starting_row, col)
            cell_value = str(cur_cell.value)
            if cell_value not in columns: 
                columns.append(str(cur_cell.value))
            else:
                col_count = 1
                col_name = str(cell_value) + " ({})".format(col_count)
                while col_name in columns:
                    col_count += 1
                    col_name = str(cell_value) + " ({})".format(col_count)
                columns.append(col_name)
                cell_value = col_name
            attribute_dict[cell_value] = {}
            
            column = sheet.column_dimensions[get_column_letter(col)]
            attribute_dict[cell_value]['width'] = column.width
            attribute_dict[cell_value]['number_format'] = cur_cell.number_format
            attribute_dict[cell_value]['alignment'] = copy(cur_cell.alignment)

        old_file.close()
        
        old_df['version'] = 'old'
        new_df['version'] = 'new'

        self.terminal.print_terminal('Comparing Entries...')
        # Get each unique project
        old_projects_all = set(old_df[self.pivot_column])
        new_projects_all = set(new_df[self.pivot_column])

        deleted_projects = old_projects_all-new_projects_all
        new_projects = new_projects_all-old_projects_all

        all_data = pd.concat([old_df,new_df])
        changes = all_data.drop_duplicates(subset=columns,keep=False)

        dupe_projs = changes[changes[self.pivot_column].duplicated()==True][self.pivot_column].tolist()
        dupes = changes[changes[self.pivot_column].isin(dupe_projs)]

        # Pull out the old and new data into separate dataframes
        change_new = dupes[(dupes["version"] == "new")]
        change_old = dupes[(dupes["version"] == "old")]

        # Drop the temp columns - we don't need them now
        change_new = change_new.drop(['version'], axis=1)
        change_old = change_old.drop(['version'], axis=1)

        # Index on the account numbers
        change_new.set_index(self.pivot_column, inplace=True)
        change_old.set_index(self.pivot_column, inplace=True)

        # Combine all the changes together
        df_all_changes = pd.concat([change_old, change_new],
                                    axis='columns',
                                    keys=['old', 'new'],
                                    join='outer')
        df_all_changes = df_all_changes.swaplevel(axis='columns')[change_new.columns[0:]]
        df_changed = df_all_changes.groupby(level=0, axis=1).apply(
            lambda frame: frame.apply(ExcelComparer.report_diff, axis=1))
        df_changed = df_changed.reset_index()

        # Sets Removed and Added Entry Dataframes
        df_removed = changes[changes[self.pivot_column].isin(deleted_projects)]
        df_added = changes[changes[self.pivot_column].isin(new_projects)]
        self.terminal.print_terminal('Exporting Calculations...')
        writer = pd.ExcelWriter("my-diff.xlsx")
        
        # Writes dataframes to excel file
        df_changed.to_excel(writer,"changed", index=False, columns=columns)
        df_removed.to_excel(writer,"removed",index=False, columns=columns)
        df_added.to_excel(writer,"added",index=False, columns=columns)

        # Initializes Pandas Workbook Variables
        workbook = writer.book
        changed_sheet = writer.sheets['changed']
        removed_sheet = writer.sheets['removed']
        added_sheet = writer.sheets['added']
        c_max_row, c_max_col = df_changed.shape
        r_max_row, r_max_col = df_removed.shape
        a_max_row, a_max_col = df_added.shape

        # Adds tables to excel file
        column_settings = [{'header':column} for column in columns]
        changed_sheet.add_table(0,0, max(c_max_row,1), c_max_col-1, {'columns':column_settings, 'style':'Table Style Medium 9'})
        removed_sheet.add_table(0,0, max(r_max_row,1), r_max_col-1, {'columns':column_settings, 'style':'Table Style Medium 10'})
        added_sheet.add_table(0,0, max(a_max_row,1), a_max_col-1, {'columns':column_settings, 'style':'Table Style Medium 11'})
        writer.save()
        
        # Opens excel file with openpyxl module
        self.terminal.print_terminal('Formating Tables...')
        workbook = openpyxl.load_workbook("my-diff.xlsx")
        
        # Creates the same format as original file
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            for col in range(1, sheet.max_column+1):
                cur_cell = str(sheet.cell(row=1,column=col).value)
                
                column = sheet.column_dimensions[get_column_letter(col)]
                if cur_cell not in attribute_dict: continue
                column.width = attribute_dict[cur_cell]['width'] 
                for row in range(1, sheet.max_row+1):
                    sheet.cell(row=row, column=col).number_format = attribute_dict[cur_cell]['number_format']
                    sheet.cell(row=row, column=col).alignment = attribute_dict[cur_cell]['alignment']

        # Returns completed excel file
        workbook.save("my-diff.xlsx")
        workbook.close()
        self.terminal.print_terminal('Complete!')
        os.startfile("my-diff.xlsx")
        self.terminal.hide_terminal()
        return True

    def report_diff(x):
        if str(x[0]) == str(x[1]):
            return x[0]
        #else: 
        #    return "PREVIOUS:\n{}\n\nCHANGED:\n{}".format(str(x[0]), str(x[1]))
        if '\n' not in str(x[1]): 
            return "CHANGED: \n{}".format(x[1])
        x0_list = str(x[0]).split('\n')
        x1_list = str(x[1]).split('\n')

        oldval = ""
        changedval = "\n\nCHANGED:\n"
        removedval = "\n\nREMOVED: \n"
        for xval in x1_list:
            if xval in x0_list:
                oldval += xval + "\n"
            else: 
                changedval += xval+"\n"

        for xval in x0_list:
            if xval not in x1_list:
                removedval += xval+"\n"
        returnval = oldval+changedval
        if removedval != "\n\nREMOVED: \n":
            returnval += removedval 
        return returnval.strip('\n').strip()

if __name__ == '__main__':
    root = tk.Tk()
    width = 1280
    height = 720
    root.minsize(width, height)
    root.geometry('+{}+{}'.format(
        int((root.winfo_screenwidth()-width)/2),
        int((root.winfo_screenheight()-height)/2)))
    program = ExcelComparer(root)
    program.show_window()
    tk.mainloop()
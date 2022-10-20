import os
from signal import strsignal
import time
import cv2
import xlrd
import PyPDF2
import shutil
import openpyxl
import imghdr
import threading
import numpy as np
import tkinter as tk
from PIL import Image
from pdf2image import convert_from_path
from PyPDF2 import PdfFileMerger, PdfFileReader
from tkinter import filedialog, ttk ,NORMAL, DISABLED, WORD, RIGHT,Y, CENTER, END

from Backend.exports import export_init
from Misc.module_imports import POPPLER

PROGRAMTITLE = 'Drawing List Organizer'
PROGRAMFOLDER = export_init(PROGRAMTITLE)

INPUTFOLDER = os.path.join(PROGRAMFOLDER, 'input')
if not os.path.exists(INPUTFOLDER): os.mkdir(INPUTFOLDER)
OUTFOLDER = os.path.join(PROGRAMFOLDER, 'output')
if not os.path.exists(OUTFOLDER): os.mkdir(OUTFOLDER)
MATCHEDFOLDER = os.path.join(OUTFOLDER, 'Matched Files')
if not os.path.exists(MATCHEDFOLDER): os.mkdir(MATCHEDFOLDER)

class DLOrganize():
    programtitle='Drawing List Organizer'
    description="""\t\t     %s
            \n\nThis program organizes a set of drawings for printing and project binding.
            \n\nA drawing list is required for to use this program.
            \n\n-------------------------------------------------------------------------------------------
            \n\nFor more help with how this program works, please contact the Program Designer for assistance.
            \n\nProgram Designer: Ethan de Leon
            \n\nRelease Date: September 28, 2021
            \n\nThis software was made for the purposes of Ocampo-Esta Corporation only. Any usesage of this program outside of the company workplace is strictly prohibited.
            """%programtitle
    bgcolor='orange'
    
    def __init__(self):
        master = tk.Tk()
        self.master = master

        global entryboxX,entryboxWidth,entryboxlabelx
        entryboxX=90
        entryboxlabelx=12
        entryboxWidth=75
        padding=5

        self.frame = tk.Frame(self.master, bg = self.bgcolor)
        self.frame.configure(bg=self.bgcolor)
        self.frame.pack()

        self.titlelabel = tk.Label(self.frame, text=self.programtitle, bg = self.bgcolor)
        self.titlelabel.config(font=('Arial Italic', 20,'bold'))
        self.titlelabel.grid(row=0,column=0, columnspan=4,pady=10)

        self.spreadSheetEntry       =tk.Entry(self.frame,width=entryboxWidth)
        self.filesEntry             =tk.Text(self.frame,width=56, height=10)

        self.sheetnameLabel         =tk.Label(self.frame,text="Sheetname:",bg=self.bgcolor)
        self.sheetnamevar           =tk.StringVar()
        self.sheetchoices           =("-----","None")
        self.sheetnameEntry         =ttk.OptionMenu(self.frame,self.sheetnamevar, *self.sheetchoices)

        self.startingLabel         =tk.Label(self.frame,text="Starting Row:",bg=self.bgcolor)
        self.startingEntry         =tk.Entry(self.frame,width=entryboxWidth)

        self.spreadSheetButton      =tk.Button(self.frame,cursor='hand2', text="Spread Sheet*:", command=lambda m=master: self.askSpreadSheet(m),bg='lemon chiffon')
        self.filesButton            =tk.Button(self.frame,cursor='hand2', text="       Files*:        ", command=lambda m=master: self.askFiles(m),bg='lemon chiffon')
        self.startButton            =tk.Button(self.frame,cursor='hand2', text='           Start           ', command=lambda m=self.confirmStart: threading.Thread(target=m).start(),bg='green',font=('helvetica', 12, 'bold'),fg='white')

        self.var1                   =tk.IntVar()
        self.var2                   =tk.IntVar()
        self.numberMarkButton       =tk.Checkbutton(self.frame,cursor='hand2', text='Stamp Numbers',variable=self.var1, onvalue=1, offvalue=0,bg = self.bgcolor)
        self.binderButton           =tk.Checkbutton(self.frame,cursor='hand2', text='Bind Files',variable=self.var2, onvalue=1, offvalue=0,bg = self.bgcolor)

        self.rowcount=1
        self.spreadSheetButton      .grid(row=self.rowcounter(),column=0,padx=10)
        self.spreadSheetEntry       .grid(row=self.rowcount,column=1,columnspan=3)

        self.sheetnameLabel         .grid(row=self.rowcounter(),column=0,padx=10)
        self.sheetnameEntry         .grid(row=self.rowcount,column=1,columnspan=3,pady=5,sticky='w')

        self.startingLabel          .grid(row=self.rowcounter(),column=0,padx=10)
        self.startingEntry          .grid(row=self.rowcount,column=1,columnspan=3,pady=5)

        self.filesButton            .grid(row=self.rowcounter(),column=0)
        self.filesEntry             .grid(row=self.rowcount,column=1, padx=5,columnspan=3,pady=5)

        self.numberMarkButton       .grid(row=self.rowcounter(),column=1, pady=3,sticky='w')
        self.binderButton           .grid(row=self.rowcount,column=2, pady=3,sticky='w')
        self.startButton            .grid(row=self.rowcounter(),column=0,columnspan=4,pady=padding)

        self.files=None
        self.spreadsheet=None

        tk.mainloop()

    def rowcounter(self):
        self.rowcount+=1
        return self.rowcount

    def askSpreadSheet(self,master):
        self.spreadSheetEntry.configure(state=NORMAL)
        filepath = filedialog.askopenfilename(title="Please Select a Spreadsheet File", filetypes=(("All Files","*.*"),("Excel","*.xlsx"),("Excel 97","*.xls")))
        if filepath=="": return False

        self.spreadSheetEntry.delete(0,'end')
        self.spreadsheet=filepath            
        self.spreadSheetEntry.insert(0,filepath)
        self.spreadSheetEntry.configure(state=DISABLED)

        wb=openpyxl.load_workbook(self.spreadsheet)
        self.sheetchoices=wb.sheetnames

        self.sheetnamevar.set(self.sheetchoices[0])
        self.sheetnameEntry['menu'].delete(0,'end')
        for choice in self.sheetchoices:
            self.sheetnameEntry['menu'].add_command(label=choice,command=tk._setit(self.sheetnamevar,choice))

    def askFiles(self,master):
        self.filesEntry.configure(state=NORMAL)
        filepath = filedialog.askopenfilenames(title="Please Select the Associated Files")
        if filepath!="":
            self.filesEntry.delete('1.0','end')
            self.files=filepath
            self.filesEntry.insert('1.0',filepath)
            self.filesEntry.configure(state=DISABLED)
        if self.filesEntry.get('1.0',END)!="":
            self.filesEntry.configure(state=DISABLED)

    def confirmStart(self):
        # region Initialization
        # retrieve inputs
        sheetname=self.sheetnamevar.get()
        var1=self.var1.get()
        var2=self.var2.get()
        self.variable1=var1

        self.outputfiles=[]
        # Finding the header and column titles
        try: 
            rrow=int(self.startingEntry.get())
        except:
            rrow=1

        from widgets.progressbar import PBar
        self.p_window=PBar(title ='File Organizer',initialwarn='Do Not Close Or Exit This Window')

        # Valid input check
        if (self.files==None)or(self.spreadsheet==None)or(self.filesEntry.get=="")or(self.spreadSheetEntry.get==""):
            msg=tk.messagebox.showerror('Error','Required file(s) not found')
            return
        
        # Deletes existing files in the Input Folder and the Matched Folder
        if len(os.listdir(INPUTFOLDER))>0:
            for f in os.listdir(INPUTFOLDER): os.remove(os.path.join(INPUTFOLDER,f))
        if len(os.listdir(MATCHEDFOLDER))>0:
            for f in os.listdir(MATCHEDFOLDER): os.remove(os.path.join(MATCHEDFOLDER,f))
        
        # endregion
        
        # region Spreadsheet Reading

        drawingcolumn=None     #int(drawingcolumn)
        titlecolumn=None
        startingPos=None       #int(startingPos)

        # Converts .xls files to .xlsx
        if self.spreadsheet.endswith('xls'): self.spreadsheet=DLOrganize.open_xls_as_xlsx(self.spreadsheet,INPUTFOLDER)

        wb=openpyxl.load_workbook(self.spreadsheet)
        if (sheetname=="")or(sheetname not in wb.sheetnames): sheet=wb.active
        else: sheet=wb[sheetname]
        
        header_row=None
        while (rrow<=sheet.max_row)and(header_row==None):
            col=1
            while col<=sheet.max_column:
                cell=str(sheet.cell(column=col,row=rrow).value)
                col+=1
            
                if "DRAWING" in cell.upper()and(drawingcolumn==None):
                    header_row=rrow
                    drawingcolumn=col-1
                    continue
                elif "DWG" in cell.upper()and(drawingcolumn==None):
                    header_row=rrow
                    drawingcolumn=col-1
                    continue

                if ("TITLE" in cell.upper())and(titlecolumn==None):
                    header_row=rrow
                    titlecolumn=col-1
                    continue
                elif("DESCRIPTION" in cell.upper())and(titlecolumn==None):
                    header_row=rrow
                    titlecolumn=col-1
                    continue

            rrow+=1

        if (drawingcolumn==None):
            msg=tk.messagebox.showerror('Drawing # Column Not Found','Drawing List does not have a drawing # column with the column title containing \'Drawing\'')
            return

        self.startButton.configure(text='In Progress',bg='grey90',command=None)
        self.p_window.warningLabel.configure(text='Do Not Close or Exit This Window.')
        self.p_window.p1Label.configure(text='Reading Files...')
        self.p_window.step(10)

        headerdrawing=str(sheet.cell(column=drawingcolumn,row=header_row).value)
        indexer=1
        startingPos=header_row+indexer
        cellCount=0
        drawinglist=[]
        titlelist=[]

        #reads excel file and gets all drawing numbers
        while (startingPos+cellCount<=sheet.max_row):
            drawingnum=sheet.cell(column=drawingcolumn,row=startingPos+cellCount).value
            if (drawingnum==None):
                cellCount+=1
                continue
            drawingnum=str(drawingnum)
            drawingnum=drawingnum.replace('O','0')
            if drawingnum==headerdrawing:
                cellCount+=1
                continue
            drawinglist.append(drawingnum)

            if titlecolumn!=None:
                title=sheet.cell(column=titlecolumn,row=startingPos+cellCount).value
                title=str(title)
                titlelist.append(title)
            cellCount+=1

        self.p_window.p1Label.configure(text='Matching Files...')
        self.p_window.step(10)
        #creates a matching list and a missing list
        matchlist=[]
        missedlist=[]
        for d in range(len(drawinglist)):
            matchnum=len(matchlist)
            for f in self.files:
                if not (os.path.isfile(f)):
                    continue
                
                if str(drawinglist[d]) in f:
                    if titlecolumn!=None:
                        matchlist.append((d+1,drawinglist[d],f,titlelist[d]))
                    else:
                        matchlist.append((d+1,drawinglist[d],f))
            #if there wasn't a file found, then the entry will be added to the missed entries list
            if len(matchlist)==matchnum:
                if titlecolumn!=None:
                    missedlist.append((d+1,drawinglist[d],titlelist[d]))
                else:
                    missedlist.append((d+1,drawinglist[d]))
            
        
        #creates list of files that were not needed in organization
        missedfiles=[]
        for file in self.files:
            matched=False
            for match in matchlist:
                if file==match[2]: matched=True
            if not matched: missedfiles.append(file)

        #reorders the matching list
        #sortedMatchlist=[]
        #endfiles=[]
        #for match in matchlist:
        #    if match[1][0]=='0':
        #        endfiles.append(match)
        #    else:
        #        sortedMatchlist.append(match)
        #matchlist=sortedMatchlist+endfiles

        self.matchlist=matchlist
        self.missedfiles=missedfiles
        self.missedlist=missedlist
        self.filelist=[]

        #creates a list of all matched file locations and copies files into the output.Matched Files directory.
        for match in matchlist:
            self.filelist.append(match[2])
            basename = os.path.basename(match[2])
            ext = os.path.splitext(match[2])[1]
            basename = basename.replace(ext, '')
            try: 
                titlename = ', '+file_clean_str(match[3])
            except:
                titlename = ''
            filename=f"{str(matchlist.index(match)+1)}-{basename}{titlename}{ext}"
            filecopy=os.path.join(MATCHEDFOLDER,filename)
            shutil.copy(match[2],filecopy)
            self.p_window.step(40/len(matchlist))
            time.sleep(1)

        #adds img number to images
        if var1==1:
            self.p_window.p1Label.configure(text='Stamping Numbers...')
            self.numberStamping()

        #File Binding
        if (var2==1):
            self.p_window.p1Label.configure(text='Binding Files...')
            self.fileBinding3(self.filelist)
            if (var1==1):
                self.p_window.p1Label.configure(text='Binding Marked Files...')
                self.fileBinding3(self.pdflistMarked,savename='Drawing Binder_M.pdf')

        #Lists results of the organization
        printtext="Matched Drawings: \n"
        if len(matchlist)>0:
            for m in range(len(matchlist)):
                if titlecolumn==None:
                    printtext+=(str(matchlist[m][0])+". "+str(matchlist[m][1])+'\n'+str(matchlist[m][2])+'\n\n')
                else:
                    printtext+=(str(matchlist[m][0])+". "+str(matchlist[m][1])+'\n'+str(matchlist[m][3])+'\n'+str(matchlist[m][2])+'\n\n')
            printtext+='\r'
        else:
            printtext+='None'+'\n\n\r'
        missedEntriestext="Missed Drawings: \n"
        if len(missedlist)>0:
            for n in range(len(missedlist)):
                if titlecolumn==None:
                    missedEntriestext+=str(missedlist[n][0])+". "+str(missedlist[n][1])+'\n'
                else:
                    missedEntriestext+=str(missedlist[n][0])+". "+str(missedlist[n][1])+'\n'+str(missedlist[n][2])+'\n\n'
            missedEntriestext+='\r'
        else:
            missedEntriestext+='None'+'\n\n\r'
    
        missedFilestext="\nMissed Files: \n"
        if len(missedfiles)>0:
            for fs in range(len(missedfiles)):
                missedFilestext+=str(missedfiles[fs])+'\n'
            missedFilestext+='\r'
        else:
            missedFilestext+='None'+'\n\n\r'

        matchedTotal        ='Total Matched Entries: '+str(len(matchlist))+'\n\n'
        missedEntriesTotal  ='Total Missed Entries: '+str(len(missedlist))+'\n\n'
        missedFilesTotal    ='Files Not Matched: '+str(len(missedfiles))+'\n\n'
        
        totaltext = matchedTotal + missedEntriesTotal + missedFilesTotal
        resulttext= totaltext + printtext + missedEntriestext + missedFilestext

        #open text file
        resultsfile=os.path.join(OUTFOLDER,'Results.txt')
        text_file = open(resultsfile, "w")
        #write string to file
        text_file.write(resulttext)
        #close file
        text_file.close()
        self.outputfiles.append(resultsfile)

        #deletes all files that were not created in this process
        for q in os.listdir(OUTFOLDER):
            if not os.path.isfile(os.path.join(OUTFOLDER,q)):
                continue
            if os.path.join(OUTFOLDER,q) not in self.outputfiles:
                os.remove(os.path.join(OUTFOLDER,q))

        self.p_window.step(100)
        time.sleep(2)
        self.p_window.destroy()
        
        #Displays results
        self.resultwindow('File Organizer Results',resulttext,self.bgcolor)
        

        #Clears entries and progress status widgets
        try:
            self.entryClear()
            self.startButton.configure(text='           Start           ', command=lambda m=self.confirmStart: threading.Thread(target=m).start(),bg='green')
        except:
            pass
        return
    
    def entryClear(self):
        self.filesEntry.configure(state=NORMAL)
        self.spreadSheetEntry.configure(state=NORMAL)
        self.spreadSheetEntry.delete(0,'end')
        self.filesEntry.delete('1.0','end')

    def askprint(self):
        #msg=tk.messagebox.askyesno('Print?','Would you like to print the Results?')
        #if msg:
        #    Printing(self.pdflist,origin=self.resultswindow)
        #    return
        self.resultswindow.destroy()

    def numberStamping(self):
        imglist=[]
        pdflist=[]
        keypoint_font=cv2.FONT_HERSHEY_SIMPLEX
        scale=3.5
        bgr=(255,150,10)
        for match in self.matchlist:
            filename=os.path.basename(match[2])
            self.p_window.p2Label.configure(text='(%d/%d) Stamping: %s'%(self.matchlist.index(match)+1,len(self.matchlist), filename))
            split_path=os.path.splitext(match[2])
            file_ext=split_path[1]
            file_ext=file_ext.lower()
            img_file=filename
            img_file=img_file.replace(file_ext,"")
            img_file=img_file.replace(file_ext.upper(),"")
            jpeg_file=img_file+".jpg"

            if match[0]>=100:
                numLoc=200
            elif match[0]>=10:
                numLoc=160
            else:
                numLoc=120

            if file_ext=='.pdf':
                images = convert_from_path(match[2],poppler_path=POPPLER,fmt='jpg')
                imagelist=[]
                for i in range(len(images)):
                    image=images[i]
                    if i!=0:
                        jpeg_file=img_file+'-'+str(i)+'.jpg'
                    jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                    image.save(jpgLoc,'JPEG')
                    
                    if i==0:
                        image=cv2.imread(jpgLoc)
                        height, width, _ = image.shape
                        shapes=np.zeros_like(image, np.uint8)
                        cv2.putText(shapes,str(match[0]),(width-numLoc,height-40),keypoint_font,scale,bgr,thickness=3)
                        cv2.circle(shapes, (width-80,height-80), 80, bgr, thickness=3)
                        outputImg=image.copy()
                        alpha=.5
                        mask=shapes.astype(bool)
                        outputImg[mask]=cv2.addWeighted(image, alpha, shapes, 1 - alpha, 0)[mask]

                        cv2.imwrite(jpgLoc,outputImg)
                    imglist.append(jpgLoc)
                    imagelist.append(Image.open(jpgLoc))
                pdfpath=os.path.join(OUTFOLDER,img_file+'_M.pdf')
                imagelist[0].save(pdfpath,'PDF',save_all=True,append_images=imagelist[1:])
                pdflist.append(pdfpath)
            else:
                if imghdr.what(match[2])==None:
                    continue
                jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                image=Image.open(match[2])
                image = image.convert('RGB')
                image.save(jpgLoc,'JPEG')

                image=cv2.imread(jpgLoc)
                height, width, _ = image.shape
                shapes=np.zeros_like(image, np.uint8)
                cv2.putText(shapes,str(match[0]),(width-numLoc,height-40),keypoint_font,scale,bgr,thickness=3)
                cv2.circle(shapes, (width-80,height-80), 80, bgr, thickness=3)
                outputImg=image.copy()
                alpha=.5
                mask=shapes.astype(bool)
                outputImg[mask]=cv2.addWeighted(image, alpha, shapes, 1 - alpha, 0)[mask]

                cv2.imwrite(jpgLoc,outputImg)
                imglist.append(jpgLoc)
                pdfpath=os.path.join(OUTFOLDER,img_file+'_M.pdf')
                image=Image.open(jpgLoc)
                image.save(pdfpath,'PDF')
                pdflist.append(pdfpath)
            
            self.p_window.step(20/len(self.matchlist))

        self.pdflistMarked=pdflist
        self.outputfiles=self.outputfiles+pdflist

        #deletes all temporary images
        for img in imglist:
            os.remove(img)
        self.p_window.p2Label.configure(text='')
        return self.pdflistMarked

    # region File Binding
    def fileBinding(self, filelist):
        merger=PdfFileMerger()

        for file in filelist:
            filename=os.path.basename(file)
            split_path=os.path.splitext(file)
            file_ext=split_path[1]
            file_ext=file_ext.lower()
            if file_ext=='.pdf':
                merger.append(PdfFileReader(file,'rb'))
            else:
                img_file=filename
                img_file=img_file.replace(file_ext,"")
                img_file=img_file.replace(file_ext.upper(),"")
                pdfpath=os.path.join(PROGRAMFOLDER,img_file+'.pdf')
                image=Image.open(file)
                image.save(pdfpath,'PDF')
                merger.append(PdfFileReader(pdfpath,'rb'))

        bindername='Drawing Binder.pdf'
        binder=os.path.join(OUTFOLDER,bindername)
        self.outputfiles.append(binder)
        merger.write(binder)
        merger.close()

    def fileBinding2(self,filelist):
        imglist=[]        
        for file in filelist:
            filename=os.path.basename(file)
            self.p_window.p2Label.configure(text='(%d/%d) Binding: %s'%(filelist.index(file)+1,len(filelist), filename))
            split_path=os.path.splitext(file)
            file_ext=split_path[1]
            file_ext=file_ext.lower()
            img_file=filename
            img_file=img_file.replace(file_ext,"")
            img_file=img_file.replace(file_ext.upper(),"")
            jpeg_file=img_file+".jpg"
            if file.endswith('pdf'):
                images = convert_from_path(file,poppler_path=POPPLER,fmt='jpg')
                for i in range(len(images)):
                    image=images[i]
                    if i!=0:
                        jpeg_file=img_file+'-'+str(i)+'.jpg'
                    jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                    image.save(jpgLoc,'JPEG')
                    imglist.append(jpgLoc)
            else:
                if imghdr.what(file)==None:
                    continue
                jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                image=Image.open(file)
                image = image.convert('RGB')
                image.save(jpgLoc,'JPEG')
                imglist.append(jpgLoc)
        
        openimagelist=[]
        for img in imglist:
            openimagelist.append(Image.open(img))
        savepath=os.path.join(OUTFOLDER,'Drawing Binder.pdf')
        openimagelist[0].save(savepath,'PDF',save_all=True,append_images=openimagelist[1:])
        self.outputfiles.append(savepath)
        
        #deletes all temporary images
        for img in imglist:
            os.remove(img)
        self.p_window.p2Label.configure(text='')
        
    def fileBinding3(self,filelist,savename='Drawing Binder.pdf'):
        merger=PdfFileMerger(strict=False)
        for file in filelist:
            filename=os.path.basename(file)
            self.p_window.p2Label.configure(text='(%d/%d) Binding: %s'%(filelist.index(file)+1,len(filelist), filename))
            try:
                with open(file,'rb') as pdf_file:
                    merger.append(PdfFileReader(pdf_file))
            
            #if the pdf cannot be read properly
            except PyPDF2.utils.PdfReadError:
                self.p_window.p2Label.configure(text='(%d/%d) Converting: %s'%(filelist.index(file)+1,len(filelist), filename))
                #converts file into pdf to read
                split_path=os.path.splitext(file)
                file_ext=split_path[1]
                img_file=filename
                img_file=img_file.replace(file_ext,"")
                file_ext=file_ext.lower()
                if file.endswith('pdf'):
                    images = convert_from_path(file,poppler_path=POPPLER,fmt='jpg')
                    imagelist=[]
                    for i in range(len(images)):
                        image=images[i]
                        jpeg_file=img_file+'-'+str(i+1)+'.jpg'
                        jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                        image.save(jpgLoc,'JPEG')
                        imagelist.append(Image.open(jpgLoc))
                    pdfpath=os.path.join(OUTFOLDER,img_file+'_M.pdf')
                    imagelist[0].save(pdfpath,'PDF',save_all=True,append_images=imagelist[1:])
                else:
                    if imghdr.what(file)==None:
                        continue
                    jpeg_file=img_file+".jpg"
                    jpgLoc=os.path.join(PROGRAMFOLDER,jpeg_file)
                    image=Image.open(file)
                    image = image.convert('RGB')
                    image.save(jpgLoc,'JPEG')
                    pdfpath=os.path.join(OUTFOLDER,img_file+'_M.pdf')
                    image=Image.open(jpgLoc)
                    image.save(pdfpath,'PDF')
                
                with open(pdfpath,'rb') as pdf_file:
                    merger.append(PdfFileReader(pdf_file))
            self.p_window.step(20/len(self.filelist))
        savepath=os.path.join(OUTFOLDER,savename)
        merger.write(savepath)
        merger.close()
        self.outputfiles.append(savepath)
    # endregion

    # region Result Output
    def resultwindow(self,title,texts,bgcolor,on_closing=None):
        self.resultswindow=tk.Toplevel()
        self.resultswindow.grab_set()
        self.resultswindow.wm_title(title)
        w_width=500
        w_height=600
        self.resultswindow.geometry('%dx%d'%(w_width,w_height))
        self.resultswindow.resizable(0,0)
        bgcolorset=bgcolor
        if bgcolor=='white':
            bgcolorset='gray90'
        self.resultswindow.configure(bg=bgcolorset)

        if on_closing!=None:
            self.resultswindow.protocol("WM_DELETE_WINDOW",on_closing)

        scrollbarWidth=20
        boarder=20

        resultbox=tk.Text(self.resultswindow,font=('helvetica',12),wrap=WORD)
        resultbox.place(x=10,y=10,width=w_width-boarder-scrollbarWidth,height=w_height-100)
        resultbox.insert('1.0',texts)
        resultbox.configure(state=DISABLED)

        scrollbar=ttk.Scrollbar(self.resultswindow, orient='vertical',command=resultbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        resultbox['yscrollcommand']=scrollbar.set

        if self.variable1==1:
            printlist=self.pdflistMarked
        else:
            printlist=self.filelist
        printbutton=tk.Button(self.resultswindow, text='Print',font=('helvetica',12,'bold'))
        printbutton.place(x=100,y=w_height-50,anchor=CENTER,width=100,height=50)

        folderButton=tk.Button(self.resultswindow, text='Folder', command=lambda m=OUTFOLDER: os.startfile(m),font=('helvetica',12,'bold'))
        folderButton.place(x=380,y=w_height-50,anchor=CENTER,width=100,height=50)

    def open_xls_as_xlsx(filename,saveloc):
        # first open using xlrd
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        nrows = sheet.nrows
        ncols = sheet.ncols

        # prepare a xlsx sheet
        book1 = openpyxl.Workbook()
        sheet1 = book1.active
        for row in range(1,nrows):
            for col in range(1,ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row-1, col-1)
        filesplit=os.path.splitext(os.path.basename(filename))

        filesave=os.path.join(saveloc,filesplit[0]+'.xlsx')
        book1.save(filename=filesave)

        return filesave

    # endregion

def file_clean_str(strs):
    bad_chars = ['#','%','&','{','}','\\','/','!',"'",'"',':','@','<','>',
        '*','?','+','`','|','=']
    new_str = strs
    for c in bad_chars:
        new_str = new_str.replace(c,'')
    return new_str.strip()


def main():
    program = DLOrganize()

if __name__=='__main__':
    main()
